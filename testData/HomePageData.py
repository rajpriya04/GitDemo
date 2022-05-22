from testData.excelDemo import sheet
import openpyxl

class HomePageData:

    test_homePage_data = [
        {"name": "Priya", "email": 'raj.rpriya90@gmail.com', "password": 'love', "gender1": "Female", "status": 'Option1'},
        {"name": "Raj", "email": 'raj.rpriya04@gmail.com', "password": 'blind', "gender1": "Male", "status": 'Option2'}]

    @staticmethod
    def getExcelData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Priya uday\\Documents\\Python\\PythonDemo.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)


        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    # print(sheet.cell(row=i, column=j).value)

        return[Dict]